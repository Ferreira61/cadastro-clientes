import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook



# Setando a aparencia padrão
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.all_system()

    def layout_config(self):
        self.title("Sistema de Cadastro de Clientes")
        self.geometry("700x500")
        self.resizable(False, False)

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=["#000", "#fff"]).place(x=50, y=390) #Para ligth e Dark
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Dark", "Light", "System"], command=self.change_apm).place(x=50, y=420)
        


    def all_system(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="SteelBlue")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Cadastro de Clientes", font=("Century Gothic bold", 22), text_color="#fff").place(x=190, y=10)

        span = ctk.CTkLabel(self, text="Preencha todos os campos do formulário para realizar o cadastro!", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)

        ficheiro = pathlib.Path("Clientes.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro=Workbook()
            folha=ficheiro.active
            folha['A1']="Nome"
            folha['B1']="Contato"
            folha['C1']="Idade"
            folha['D1']="Gênero"
            folha['E1']="Email"
            folha['F1']="Endereço"
            folha['G1']="Estado"
            folha['H1']="CEP"

            ficheiro.save("Clientes.xlsx")

        
        def submit():
            
            #Pegando dados dos entrys
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            email = email_value.get()
            adress = adress_value.get()
            state = state_combobox.get()
            cep = cep_value.get()

            if (name=="" or contact=="" or age=="" or email=="" or state=="" or cep==""):
                messagebox.showerror("Sistema", "ERRO\nPreencha os dados pendentes!")
            else:
                
              ficheiro = openpyxl.load_workbook('Clientes.xlsx')
              folha = ficheiro.active
              folha.cell(column=1, row=folha.max_row+1, value=name)
              folha.cell(column=2, row=folha.max_row, value=contact)
              folha.cell(column=3, row=folha.max_row, value=age)
              folha.cell(column=4, row=folha.max_row, value=gender)
              folha.cell(column=5, row=folha.max_row, value=email)
              folha.cell(column=6, row=folha.max_row, value=adress)
              folha.cell(column=7, row=folha.max_row, value=state)
              folha.cell(column=8, row=folha.max_row, value=cep)

              ficheiro.save(r"Clientes.xlsx")
              messagebox.showinfo("Sistema", "Dados enviados e salvos com sucesso!") 


        def clear():
            #Apagando dados dos entrys
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            email_value.set("")
            adress_value.set("")
            cep_value.set("")



        #Texts Variables
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        email_value = StringVar()
        adress_value = StringVar()
        cep_value = StringVar()

        #Entrys
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gohtic bold", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=("Century Gohtic bold", 16), fg_color="transparent")
        email_entry = ctk.CTkEntry(self, width=200, textvariable=email_value, font=("Century Gohtic bold", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century Gohtic bold", 16), fg_color="transparent")
        adress_entry = ctk.CTkEntry(self, width=200, textvariable=adress_value, font=("Century Gohtic bold", 16), fg_color="transparent")
        cep_entry = ctk.CTkEntry(self, width=150, textvariable=cep_value, font=("Century Gohtic bold", 16), fg_color="transparent")

        #ComboBox
        gender_combobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino"], font=("Century Gothic bold", 14), width=170, state="readonly")
        gender_combobox.set("Selecione o gênero",)
        state_combobox = ctk.CTkComboBox(self, values=["AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"], font=("Century Gothic bold", 14), width=170, state="readonly")
        state_combobox.set("Selecione o estado")

        #Entry of Obs
        ##border_width=2, fg_color="transparent")

        #Labels
        lb_name = ctk.CTkLabel(self, text="Nome Completo:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(self, text="Contato:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_age = ctk.CTkLabel(self, text="Idade:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_gender = ctk.CTkLabel(self, text="Gênero:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_email = ctk.CTkLabel(self, text="Email:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_adress = ctk.CTkLabel(self, text="Endereço:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_cep = ctk.CTkLabel(self, text="CEP:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_state = ctk.CTkLabel(self, text="Estado:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=280, y=420)
        btn_clear = ctk.CTkButton(self, text="Limpar campos".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=420)

        
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)

        lb_gender.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)

        lb_email.place(x=50, y=190)
        email_entry.place(x=50, y=220)

        lb_adress.place(x=50, y=260)
        adress_entry.place(x=50, y=290)

        lb_cep.place(x=300, y=260)
        cep_entry.place(x=300, y=290)

        lb_state.place(x=500, y=260)
        state_combobox.place(x=500, y=290)

    def change_apm(self, new_appearance):
        ctk.set_appearance_mode(new_appearance)


if __name__ =="__main__":
    app = App()
    app.mainloop()
